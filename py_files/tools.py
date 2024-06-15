from openpyxl import Workbook, load_workbook
import sys, os, shutil, filecmp
from py_files.CsvFormats import format0
from py_files.Person import Person
from py_files.DataType1 import DataType1

# list of data types that exist
dataTypes = [DataType1]


def mapFiles(root, verbose = False, directory = None):
    """
    Maps files inside a directory to a dictionary

    :param root: string, path of highest level in directory you wish to map
    :param verbose: boolean set to true for output on ignored files
    :param directory: used only for recursion, do not override
    :returns: dictionary with each key as a folder which contains a list of its files
    """
    # get name of current folder
    dirname = root.split(os.path.sep)[-1]
    # check if dictionary exists and add current folder as key
    if directory == None:
        directory = {}
    if dirname != "maer_raw":
        directory[dirname] = []
    # loop through files in this directory
    for filename in os.listdir(root):
        if filename.endswith(".csv"):
            directory[dirname].append(filename)
        elif "." not in filename:
            directory = mapFiles(os.path.join(os.getcwd(), "maer_raw", filename), verbose, directory)
        else:
            # get folders starting at project folder "DataOrganizer"
            projectroot = root.split(os.path.sep)[os.getcwd().split(os.path.sep).index("DataOrganizer"):]
            if verbose:
                print("Ignoring {}".format(os.path.sep.join(projectroot)+os.path.sep+filename))
    return directory


def createDataTypes(format_id, filepath, expName = None, prt=1):
    """
    Create `DataType` objects based on the format of the input file.

    :param format: integer specifying the format identifier
    :param filepath: string of complete filepath to .csv file
    :returns: list of `DataType` objects

    """
    if format_id == 0:
        return format0(filepath, experiment=expName, prt=prt)
    else:
        print("Unsupported format: {}".format(format_id))


def addDataToPerson(listofData, listofPeople, warnOnAdd = False):
    """
    Takes data and adds it to specific people, adding new people if needed.

    It is important to save the returned data from this function as new people might be added

    :returns: list of Person objects
    """
    for data in listofData:
        person_id = data.getAssociatedParticipantID()
        for person in listofPeople:
            if person_id == person.getID():
                person.addData(data)
                break
        else:
            if warnOnAdd:
                print("WARNING: Person added: id = {}, experiment = {}".format(person_id, data.experiment))
            listofPeople.append(Person(person_id, [data]))
    return listofPeople

def createRow(person):
    """
    Create a row to publish to a workbook

    :param person: `Person` object to create a row for
    """
    person.sortData()
    ret = (person.id, person.data1[0].date, person.data1[0].experiment)
    for data in person.data1:
        ret += data.getAllFields()



def publishToWorkbook(name, rows, startCell = (1,1), verbose=False):
    """
    Publish data to workbook

    :param name: string to name file once done writing
    :param rows: list containing a list of the data for each row
    :param startCell: tuple of two numbers >= 1
    :param verbose: boolean
    :returns: tuple of two numbers representing the last cell written
    """
    if not name.endswith(".xlsx"):
        name = name.split(".")[0]
        name += ".xlsx"

    # if append:
    #     wb = load_workbook(name)
    # else:
    #     wb = Workbook()
    wb = Workbook()
    sheet = wb.active
    r = startCell[0]
    c = startCell[1]

    for row in rows:
        for item in row:
            sheet.cell(row=r, column=c).value = item
            saved_c = c
            # sheet.append(row)
            c += 1
        r += 1
        c = startCell[1]
    c = saved_c

    wb.save(name)
    if verbose:
        print("Data saved to {}".format(name))

    return (r,c)


def sublistSort(sub_li):
    # taken from https://www.geeksforgeeks.org/python-sort-list-according-second-element-sublist/
    l = len(sub_li)
    for i in range(0, l):
        for j in range(0, l - i - 1):
            if (sub_li[j] > sub_li[j + 1]):
                tempo = sub_li[j]
                sub_li[j] = sub_li[j + 1]
                sub_li[j + 1] = tempo
    return sub_li


def backup(folder_path, dest_folder = os.path.join(os.getcwd(), 'backups'), verbose = False):
    backup_files(folder_path, bakdir_name=dest_folder, verbose=verbose)


MAXVERSIONS = 100
BAKFOLDER = '.bak'
def backup_files(tree_top, bakdir_name=BAKFOLDER, verbose = False):
    # Backup files - As published in Python Cookbook
    # by O'Reilly with some bug-fixes.

    # Credit: Anand Pillai, Tiago Henriques, Mario Ruggier
    """ Directory back up function. Takes the top-level
    directory and an optional backup folder name as arguments.
    By default the backup folder is a folder named '.bak' created
    inside the folder which is backed up. If another directory
    path is passed as value of this argument, the backup versions
    are created inside that directory instead. Maximum of
    'MAXVERSIONS' simultaneous versions can be maintained

    Example usage
    -------------

    The command
    $ python backup.py ~/programs

    will create backups of every file inside ~/programs
    inside sub-directories named '.bak' inside each folder.
    For example, the backups of files inside ~/programs will
    be found in ~/programs/.bak, the backup of files inside
    ~/programs/python in ~/programs/python/.bak etc.

    The command
    $ python backup.py ~/programs ~/backups

    will create backups of every file inside ~/backups/programs
    folder. No .bak folder is created. Instead backups of
    files in ~/programs will be inside ~/backups/programs,
    backups of files in ~/programs/python will be inside
    ~/backups/programs/python etc.

    """

    top_dir = os.path.basename(tree_top)
    tree_top += os.sep

    for dir, subdirs, files in os.walk(tree_top):

        if os.path.isabs(bakdir_name):
            relpath = dir.replace(tree_top, '')
            backup_dir = os.path.join(bakdir_name, top_dir, relpath)
        else:
            backup_dir = os.path.join(dir, bakdir_name)

        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)

        # To avoid recursing into sub-directories
        subdirs[:] = [d for d in subdirs if d != bakdir_name]
        for f in files:
            filepath = os.path.join(dir, f)
            destpath = os.path.join(backup_dir, f)
            # Check existence of previous versions
            for index in range(MAXVERSIONS):
                backup = '%s.%2.2d' % (destpath, index)
                abspath = os.path.abspath(filepath)

                if index > 0:
                    # No need to backup if file and last version
                    # are identical
                    old_backup = '%s.%2.2d' % (destpath, index - 1)
                    if not os.path.exists(old_backup): break
                    abspath = os.path.abspath(old_backup)

                    try:
                        if os.path.isfile(abspath) and filecmp.cmp(abspath, filepath, shallow=False):
                            continue
                    except OSError:
                        pass

                try:
                    if not os.path.exists(backup):
                        if verbose:
                            print('Copying %s to %s...' % (filepath, backup))
                        shutil.copy(filepath, backup)
                except:
                    pass


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Usage: %s [directory] [backup directory]" % sys.argv[0])

    tree_top = os.path.abspath(os.path.expanduser(os.path.expandvars(sys.argv[1])))

    if len(sys.argv) >= 3:
        bakfolder = os.path.abspath(os.path.expanduser(os.path.expandvars(sys.argv[2])))
    else:
        bakfolder = BAKFOLDER

    if os.path.isdir(tree_top):
        backup_files(tree_top, bakfolder)